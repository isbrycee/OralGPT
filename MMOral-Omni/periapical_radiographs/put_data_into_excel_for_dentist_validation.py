import os
import json
from PIL import Image as PIL_Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XL_Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openai import OpenAI
from tqdm import tqdm

client = OpenAI(
    api_key="sk-",  # 替换成你的 DMXapi 令牌key
    base_url="https://www.dmxapi.cn/v1",  # 需要改成DMXAPI的中转 https://www.dmxapi.cn/v1 ，这是已经改好的。
)

# client = OpenAI(
#     api_key="empty",  # 替换成你的 DMXapi 令牌key
#     base_url="http://localhost:8080/v1",  # 需要改成DMXAPI的中转 https://www.dmxapi.cn/v1 ，这是已经改好的。
# )


def extract_json_to_excel_with_thumbnails(json_file: str, output_excel: str, thumbnail_size=(100, 100)):
    """
    从 JSON 文件夹提取数据并在 Excel 中显示图像缩略图
    
    参数:
        json_folder (str): JSON 文件所在文件夹路径
        output_excel (str): 输出 Excel 完整路径（需 .xlsx 后缀）
        thumbnail_size (tuple): 缩略图尺寸（宽, 高），默认 (100, 100) 像素
    """
    # 初始化数据存储列表（包含文本信息和图像路径）
    data_records = []
    
    # 遍历所有 JSON 文件
    
    try:
        # 读取 JSON 文件
        with open(json_file, 'r', encoding='utf-8') as f:
            json_content = json.load(f)
        
        if not isinstance(json_content, list):
            print(f"警告：文件 {json_file} 内容不是列表，已跳过")

        print(f"处理文件: {json_file}，包含 {len(json_content)} 条记录")
        # 遍历列表中的字典项
        for item in tqdm(json_content):
            if not isinstance(item, dict):
                print(f"警告：文件 {json_file} 中发现非字典元素，已跳过")
                continue

            # 提取关键字段（带容错）
            image_path = os.path.join("/home/jinghao/projects/x-ray-VLM/RGB/periapical_radiographs/images", item.get('image'))
            question = str(item.get('question', ''))  # 限制长度防溢出
            diagnosis = str(item.get('cot_answer', ''))
            caption = diagnosis.split('</Caption>')[0]
            answer = diagnosis.split('</Caption>')[1]

            if len(diagnosis) > 0:
                print("diagnosis:\n")
                print(answer)
                messages = [
                    {"role": "system", "content": "Help me translate the sentence into Chinese. Only output the translated Chinese, do not output anything else."},
                    {"role": "user", "content": answer},
                ]
                zn_diagnosis_res = client.chat.completions.create(
                    model="GLM-4.5-Flash",  # 或 gpt-4o
                    messages=messages,
                    temperature=0.0,
                )
                zn_diagnosis = zn_diagnosis_res.choices[0].message.content.strip().split('<Answer>')[-1].strip()
            else:
                zn_diagnosis = ""

            if len(caption) > 0:
                print("caption:\n")
                print(caption)
                messages = [
                    {"role": "system", "content": "Help me translate the sentence into Chinese. Only output the translated Chinese, do not output anything else."},
                    {"role": "user", "content": caption},
                ]
                zn_caption = client.chat.completions.create(
                    model="GLM-4.5-Flash",  # 或 gpt-4o
                    messages=messages,
                    temperature=0.0,
                )
                zn_caption = zn_caption.choices[0].message.content.strip()
                
            else:
                zn_caption = ""

            # 处理图像路径（转换为绝对路径）
            absolute_image = None
            if image_path:
                json_dir = os.path.dirname(json_file)
                absolute_image = os.path.abspath(os.path.join(json_dir, image_path))
            data_records.append({
                "原始路径": image_path,
                "绝对路径": image_path,
                "问题": question,
                "诊断英文": answer,
                "诊断中文": zn_diagnosis,
                "Caption_英文": caption,
                "Caption_中文": zn_caption,
                "Category": item.get('Disease category', []),
            })

    except json.JSONDecodeError:
        print(f"错误：文件 {json_file} 不是有效 JSON，已跳过")
    except Exception as e:
        print(f"处理文件 {json_file} 时出错：{str(e)}，已跳过")

    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "图像数据"

    # 设置表头
    headers = ["图像路径", "图像缩略图", "中文 caption", "英文 caption", "中文诊断结果", "英文诊断结果", "疾病类别", "问题描述"]
    ws.append(headers)

    # 设置列宽（根据内容调整）
    col_widths = [40, 80, 80, 80, 80, 80, 80, 50]  # 对应四列宽度（字符数）
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # 设置表头居中加粗
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for cell in ws[1]:
        cell.alignment = header_alignment
        cell.font = cell.font.copy(bold=True)

    # 遍历数据并插入 Excel（从第 2 行开始）
    for row_idx, record in enumerate(data_records, start=2):
        # 第 1 列：原始路径（字符串，非列表）
        ws.cell(row=row_idx, column=1, value=record["原始路径"])

        # 第 2 列：图像缩略图（关键逻辑，明确占位符）
        image_cell = ws.cell(row=row_idx, column=2)
        image_cell.alignment = Alignment(horizontal='center', vertical='center')

        # 检查图像路径是否有效（字符串且文件存在）
        if isinstance(record["绝对路径"], str) and os.path.exists(record["绝对路径"]):
            try:
                # 打开图像并生成缩略图
                with PIL_Image.open(record["绝对路径"]) as pil_img:
                    pil_img.thumbnail(thumbnail_size)  # 调整尺寸
                    
                    # 转换为字节流（避免临时文件）
                    img_byte_arr = io.BytesIO()
                    pil_img.save(img_byte_arr, format='PNG')  # 统一转 PNG
                    img_byte_arr.seek(0)  # 重置指针
                    
                    # 创建 Excel 图像对象并锚定到单元格
                    xl_img = XL_Image(img_byte_arr)
                    xl_img.anchor = f"B{row_idx}"  # 锚定到当前行第2列
                    ws.add_image(xl_img)
                    
            except Exception as e:
                # 图像加载失败时显示占位符（明确提示）
                image_cell.value = "图像加载失败"
                print(f"警告：图像 {record['绝对路径']} 加载失败（{str(e)}）")
        else:
            # 图像路径无效时显示占位符（明确提示）
            image_cell.value = "无有效图像路径"
            print(f"警告：图像路径无效 {record['绝对路径']}")

        # 第 3 列：问题描述
        ws.cell(row=row_idx, column=3, value=record["Caption_中文"])

        # 第 3 列：问题描述
        ws.cell(row=row_idx, column=4, value=record["Caption_英文"])

        # 第 3 列：问题描述
        ws.cell(row=row_idx, column=5, value=record["诊断中文"])

        # 第 3 列：问题描述
        ws.cell(row=row_idx, column=6, value=record["诊断英文"])

        # 第 3 列：问题描述
        ws.cell(row=row_idx, column=7, value=record["Category"])

        # 第 4 列：诊断结果
        ws.cell(row=row_idx, column=8, value=record["问题"])

    # 调整行高（适应缩略图）
    for row in ws.iter_rows(min_row=2, max_row=len(data_records)+1):
        ws.row_dimensions[row[0].row].height = thumbnail_size[1] + 20  # 增加垂直间距

    # 保存 Excel
    wb.save(output_excel)
    print(f"成功生成 Excel 文件（含缩略图）：{output_excel}")
    print(f"共处理 {len(data_records)} 条记录")

if __name__ == "__main__":
    import io  # 用于内存中处理图像字节流

    # 用户输入路径（可根据需要改为命令行参数）
    json_input = "/home/jinghao/projects/x-ray-VLM/RGB/periapical_radiographs/val.json"
    excel_output = "/home/jinghao/projects/x-ray-VLM/RGB/periapical_radiographs/test_excel_for_dentist_validation.xlsx"

    # 执行主函数（缩略图尺寸设为 100x100 像素）
    extract_json_to_excel_with_thumbnails(
        json_file=json_input,
        output_excel=excel_output,
        thumbnail_size=(500, 500)
    )