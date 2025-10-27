import json
import os
import re
import random
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
from io import BytesIO


def extract_tag_content(text, tag):
    """提取形如 <tag>内容</tag> 的内容。"""
    pattern = rf"<{tag}>(.*?)</{tag}>"
    match = re.search(pattern, text, re.DOTALL)
    return match.group(1).strip() if match else ""


def create_thumbnail(image_path, size=(400, 400)):
    """生成缩略图（返回 BytesIO 对象供 Excel 使用）。"""
    try:
        img = Image.open(image_path)
        img.thumbnail(size)
        bio = BytesIO()
        img.save(bio, format="PNG")
        bio.seek(0)
        return bio
    except Exception as e:
        print(f"⚠️ 无法生成缩略图 {image_path}: {e}")
        return None


def json_to_excel(
    json_path,
    excel_path=None,
    image_base_path=".",
    n_samples=50,
    seed=42
):
    """
    从 JSON 文件中选择包含 'cot_answer' 的 N 条样本，
    并保存到 Excel 文件（可新建或追加模式）。
    ---------------------------------------------
    :param json_path: JSON 输入文件路径
    :param excel_path: 若指定则在现有 Excel 末尾追加，否则新建 Excel
    :param image_base_path: 图片目录路径
    :param n_samples: 抽取的有效样本数量
    :param seed: 随机种子
    """
    # 1️⃣ 读取 JSON 数据
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, list):
        raise ValueError("❌ JSON 文件内容必须是一个由字典组成的列表。")

    # 2️⃣ 过滤有效样本
    valid_data = [d for d in data if "cot" in d and d["cot"].strip()]
    total_valid = len(valid_data)
    print(f"📁 过滤后共有 {total_valid} 条包含 'cot_answer' 的记录。")

    if total_valid == 0:
        raise ValueError("❌ 没有任何包含 'cot_answer' 的数据。")

    # 3️⃣ 随机抽取
    random.seed(seed)
    if n_samples > total_valid:
        n_samples = total_valid
        print(f"⚠️ 有效数据不足，只能抽取 {n_samples} 条。")

    sampled_data = random.sample(valid_data, n_samples)
    print(f"🎲 从 {total_valid} 条数据中随机抽取 {n_samples} 条（seed={seed}）。")

    # 4️⃣ 打开或创建 Excel 文件
    new_file_created = False
    if excel_path and os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
        print(f"📘 已加载现有 Excel 文件：{excel_path}")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        headers = ["image_name", "图像缩略图", "Question", "Caption", "Think", "Answer", "category"]
        ws.append(headers)
        new_file_created = True
        if excel_path is None:
            excel_path = "output.xlsx"
        print(f"🆕 创建新 Excel 文件：{excel_path}")

    # 5️⃣ 找到当前最后一行
    start_row = ws.max_row + 1

    # 6️⃣ 追加数据
    for item in sampled_data:
        file_name = item.get("file_name", "")
        question = item.get("question", "")
        cot_answer = item.get("cot", "")
        category = item.get("source", "HistopathologicalImageDiagnosis")

        caption = extract_tag_content(cot_answer, "Caption")
        think = extract_tag_content(cot_answer, "Think")
        answer = extract_tag_content(cot_answer, "Answer")

        image_path = os.path.join(image_base_path, file_name)
        thumb_io = create_thumbnail(image_path)

        row = [file_name, None, question, caption, think, answer, category]
        ws.append(row)

        # 插入图像
        if thumb_io:
            img = XLImage(thumb_io)
            img.width, img.height = 400, 400
            cell_name = f"B{ws.max_row}"
            ws.add_image(img, cell_name)
            ws.row_dimensions[ws.max_row].height = 350

    # 7️⃣ 调整列宽（如果是新建）
    if new_file_created:
        for col in ["A", "C", "D", "E", "F", "G"]:
            ws.column_dimensions[col].width = 50
    ws.column_dimensions["B"].width = 70

    # 8️⃣ 保存 Excel 文件
    wb.save(excel_path)
    print(f"✅ 已保存 Excel 文件：{excel_path} （追加了 {n_samples} 条）")


# 示例调用：
json_to_excel("/home/jinghao/projects/x-ray-VLM/RGB/MMOral-Omni/meta_json/5.1_histopathologicalImage_Diagnosis_4datasets.json", 
              excel_path="MMOral-Omni-for-human-scoring.xlsx",
              image_base_path="/home/jinghao/projects/x-ray-VLM/RGB/",
              n_samples=50, seed=123)
              