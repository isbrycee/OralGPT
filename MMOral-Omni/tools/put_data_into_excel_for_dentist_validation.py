import os
import json
from PIL import Image as PIL_Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XL_Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openai import OpenAI
from tqdm import tqdm

client = OpenAI(
    api_key="sk-xx",  # 替换成你的 DMXapi 令牌key
    base_url="https://www.dmxapi.cn/v1",  # 需要改成DMXAPI的中转 https://www.dmxapi.cn/v1 ，这是已经改好的。
)

def extract_json_to_excel_with_thumbnails(json_folder: str, output_excel: str, thumbnail_size=(100, 100)):
    """
    从 JSON 文件夹提取数据并在 Excel 中显示图像缩略图
    
    参数:
        json_folder (str): JSON 文件所在文件夹路径
        output_excel (str): 输出 Excel 完整路径（需 .xlsx 后缀）
        thumbnail_size (tuple): 缩略图尺寸（宽, 高），默认 (100, 100) 像素
    """
    # 初始化数据存储列表（包含文本信息和图像路径）
    data_records = []
    
    # 遍历所有 JSON 文件
    for root_dir, _, json_files in os.walk(json_folder):
        for json_file in json_files:
            if not json_file.lower().endswith('.json'):
                continue

            json_path = os.path.join(root_dir, json_file)
            
            try:
                # 读取 JSON 文件
                with open(json_path, 'r', encoding='utf-8') as f:
                    json_content = json.load(f)
                
                if not isinstance(json_content, list):
                    print(f"警告：文件 {json_path} 内容不是列表，已跳过")
                    continue

                print(f"处理文件: {json_path}，包含 {len(json_content)} 条记录")
                # 遍历列表中的字典项
                for item in tqdm(json_content):
                    if not isinstance(item, dict):
                        print(f"警告：文件 {json_path} 中发现非字典元素，已跳过")
                        continue

                    # 提取关键字段（带容错）
                    image_path = os.path.join("/home/jinghao/projects/x-ray-VLM/RGB/test_classification_images", item.get('image').split('OralGPT-RGB-Classification-Dataset/')[1])
                    question = str(item.get('question', ''))  # 限制长度防溢出
                    diagnosis = str(item.get('diagnosis', ''))

                    if len(diagnosis) > 0 and diagnosis != "The appearance is most consistent with a suspicious primary oral mucosal malignant melanoma involving the hard palate with extension to the adjacent maxillary alveolar/gingival mucosa." \
                        and diagnosis != "Suspicious for primary oral mucosal melanoma of the posterior hard palate extending onto the adjacent soft palate." \
                        and diagnosis != "The appearance is most suspicious for an oral mucosal malignant melanoma of the hard palate (maxillary midline/posterior palatal mucosa)." \
                        and diagnosis != "Clinical appearance is most consistent with a suspicious pigmented malignant lesion—most likely primary oral mucosal melanoma of the hard palate (midline posterior hard palate), with an associated erythematous nodular lesion of the maxillary labial gingival/frenal region suggesting possible multifocal involvement." \
                        and diagnosis != "Highly suspicious for primary oral mucosal melanoma of the hard palate with extension toward the adjacent maxillary gingiva." \
                        and diagnosis != "Suspicious for primary oral malignant melanoma of the hard palate (solitary pigmented mucosal lesion of the maxillary hard palate)." \
                        and diagnosis != "A suspicious pigmented malignant lesion most consistent with primary oral mucosal melanoma of the posterior hard palate." :
                        print(diagnosis)
                        messages = [
                            {"role": "system", "content": "Help me translate the sentence into Chinese. Only output the translated Chinese, do not output anything else."},
                            {"role": "user", "content": diagnosis},
                        ]
                        response = client.chat.completions.create(
                            model="glm-4.5-air-free",  # 或 gpt-4o
                            messages=messages,
                            temperature=0.0,
                        )
                        zn_diagnosis = response.choices[0].message.content.strip()
                    else:
                        zn_diagnosis = ""

                    # 处理图像路径（转换为绝对路径）
                    absolute_image = None
                    if image_path:
                        json_dir = os.path.dirname(json_path)
                        absolute_image = os.path.abspath(os.path.join(json_dir, image_path))
                    data_records.append({
                        "原始路径": image_path,
                        "绝对路径": image_path,
                        "问题": question,
                        "诊断英文": diagnosis,
                        "诊断中文": zn_diagnosis
                    })

            except json.JSONDecodeError:
                print(f"错误：文件 {json_path} 不是有效 JSON，已跳过")
            except Exception as e:
                print(f"处理文件 {json_path} 时出错：{str(e)}，已跳过")

    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "图像数据"

    # 设置表头
    headers = ["图像路径", "图像缩略图", "中文诊断结果", "英文诊断结果", "问题描述"]
    ws.append(headers)

    # 设置列宽（根据内容调整）
    col_widths = [40, 80, 80, 80, 50]  # 对应四列宽度（字符数）
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # 设置表头居中加粗
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for cell in ws[1]:
        cell.alignment = header_alignment
        cell.font = cell.font.copy(bold=True)

    # 遍历数据并插入 Excel（从第 2 行开始）
    for row_idx, record in enumerate(data_records, start=2):
        # 第 1 列：原始路径（字符串，非列表）
        ws.cell(row=row_idx, column=1, value=record["原始路径"])

        # 第 2 列：图像缩略图（关键逻辑，明确占位符）
        image_cell = ws.cell(row=row_idx, column=2)
        image_cell.alignment = Alignment(horizontal='center', vertical='center')

        # 检查图像路径是否有效（字符串且文件存在）
        if isinstance(record["绝对路径"], str) and os.path.exists(record["绝对路径"]):
            try:
                # 打开图像并生成缩略图
                with PIL_Image.open(record["绝对路径"]) as pil_img:
                    pil_img.thumbnail(thumbnail_size)  # 调整尺寸
                    
                    # 转换为字节流（避免临时文件）
                    img_byte_arr = io.BytesIO()
                    pil_img.save(img_byte_arr, format='PNG')  # 统一转 PNG
                    img_byte_arr.seek(0)  # 重置指针
                    
                    # 创建 Excel 图像对象并锚定到单元格
                    xl_img = XL_Image(img_byte_arr)
                    xl_img.anchor = f"B{row_idx}"  # 锚定到当前行第2列
                    ws.add_image(xl_img)
                    
            except Exception as e:
                # 图像加载失败时显示占位符（明确提示）
                image_cell.value = "图像加载失败"
                print(f"警告：图像 {record['绝对路径']} 加载失败（{str(e)}）")
        else:
            # 图像路径无效时显示占位符（明确提示）
            image_cell.value = "无有效图像路径"
            print(f"警告：图像路径无效 {record['绝对路径']}")


        # 第 3 列：问题描述
        ws.cell(row=row_idx, column=3, value=record["诊断中文"])

        # 第 3 列：问题描述
        ws.cell(row=row_idx, column=4, value=record["诊断英文"])

        # 第 4 列：诊断结果
        ws.cell(row=row_idx, column=5, value=record["问题"])

    # 调整行高（适应缩略图）
    for row in ws.iter_rows(min_row=2, max_row=len(data_records)+1):
        ws.row_dimensions[row[0].row].height = thumbnail_size[1] + 20  # 增加垂直间距

    # 保存 Excel
    wb.save(output_excel)
    print(f"成功生成 Excel 文件（含缩略图）：{output_excel}")
    print(f"共处理 {len(data_records)} 条记录")

if __name__ == "__main__":
    import io  # 用于内存中处理图像字节流

    # 用户输入路径（可根据需要改为命令行参数）
    json_input = "/home/jinghao/projects/x-ray-VLM/RGB/oral-classification-all-json-test"
    excel_output = "/home/jinghao/projects/x-ray-VLM/RGB/test_excel_for_dentist_validation.xlsx"

    # 执行主函数（缩略图尺寸设为 100x100 像素）
    extract_json_to_excel_with_thumbnails(
        json_folder=json_input,
        output_excel=excel_output,
        thumbnail_size=(500, 500)
    )