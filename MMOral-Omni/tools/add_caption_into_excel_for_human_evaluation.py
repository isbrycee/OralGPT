import os
import json
from PIL import Image as PIL_Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XL_Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openai import OpenAI
from tqdm import tqdm
from openpyxl import Workbook, load_workbook


def extract_json_to_excel_with_thumbnails(json_folder: str, output_excel: str, thumbnail_size=(100, 100)):
    """
    从 JSON 文件夹提取数据并在 Excel 中显示图像缩略图
    
    参数:
        json_folder (str): JSON 文件所在文件夹路径
        output_excel (str): 输出 Excel 完整路径（需 .xlsx 后缀）
        thumbnail_size (tuple): 缩略图尺寸（宽, 高），默认 (100, 100) 像素
    """
    # 初始化数据存储列表（包含文本信息和图像路径）
    data_records = []
    caption_mapping = {}  # 用于存储image_path到caption的映射

    # 遍历所有 JSON 文件
    for root_dir, _, json_files in os.walk(json_folder):
        for json_file in json_files:
            if not json_file.lower().endswith('.json'):
                continue

            json_path = os.path.join(root_dir, json_file)
            
            try:
                # 读取 JSON 文件
                with open(json_path, 'r', encoding='utf-8') as f:
                    json_content = json.load(f)
                
                if not isinstance(json_content, list):
                    print(f"警告：文件 {json_path} 内容不是列表，已跳过")
                    continue

                print(f"处理文件: {json_path}，包含 {len(json_content)} 条记录")
                # 遍历列表中的字典项
                for item in tqdm(json_content):
                    if not isinstance(item, dict):
                        print(f"警告：文件 {json_path} 中发现非字典元素，已跳过")
                        continue

                    # 提取关键字段（带容错）
                    cot_answer = str(item.get('cot_answer', ''))  # 限制长度防溢出
                    if len(cot_answer) == 0:
                        cot_answer = str(item.get('caption', ''))
                    caption = cot_answer.split('</caption>')[0]
                    image_path = os.path.join("/home/jinghao/projects/x-ray-VLM/RGB/test_classification_images", item.get('image').split('OralGPT-RGB-Classification-Dataset/')[1])
                    
                    
                    # 存储到映射表
                    caption_mapping[image_path] = caption
                    
                    # 添加到数据记录
                    data_records.append({
                        'image_path': image_path,
                        'caption': caption,
                        'cot_answer': cot_answer
                    })

            except json.JSONDecodeError:
                print(f"错误：文件 {json_path} 不是有效 JSON，已跳过")
            except Exception as e:
                print(f"处理文件 {json_path} 时出错：{str(e)}，已跳过")

    # 读取或创建Excel文件
    if os.path.exists(output_excel):
        wb = load_workbook(output_excel)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        # 添加表头
        ws['A1'] = 'Image Path'
        ws['B1'] = 'Caption'
        ws['C1'] = 'Original Column 2'
        ws['D1'] = 'Original Column 3'
    
    # 在第二列和第三列之间插入Caption列
    if ws.max_column >= 2:
        ws.insert_cols(3)  # 在第二列后插入新列（第三列变为新列）
        ws.cell(row=1, column=3, value='Caption')
    
    # 填充Caption数据
    for row in range(2, ws.max_row + 1):
        image_path = ws.cell(row=row, column=1).value  # 第一列是image_path
        if image_path in caption_mapping:
            ws.cell(row=row, column=3, value=caption_mapping[image_path])
    
    # 保存Excel文件
    wb.save(output_excel)
    print(f"Excel文件已更新并保存到: {output_excel}")


if __name__ == "__main__":
    import io  # 用于内存中处理图像字节流

    # 用户输入路径（可根据需要改为命令行参数）
    json_input = "/home/jinghao/projects/x-ray-VLM/RGB/oral-classification-all-json-test"
    excel_output = "/home/jinghao/projects/x-ray-VLM/RGB/test_excel_for_dentist_validation.xlsx"

    # 执行主函数（缩略图尺寸设为 100x100 像素）
    extract_json_to_excel_with_thumbnails(
        json_folder=json_input,
        output_excel=excel_output,
        thumbnail_size=(500, 500)
    )